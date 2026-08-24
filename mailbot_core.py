from __future__ import annotations

import csv
import base64
import mimetypes
import random
import re
import smtplib
import ssl
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook


PLACEHOLDER_RE = re.compile(r"\(([^()\r\n]+)\)")
EMAIL_RE = re.compile(r"^[^\s@<>]+@[^\s@<>]+\.[^\s@<>]+$")
NEWLINE_RE = re.compile(r"[\r\n]")


class MailBotError(Exception):
    """Kullanıcıya gösterilebilecek uygulama hatası."""


@dataclass(frozen=True)
class Record:
    row_number: int
    values: dict[str, str]


@dataclass(frozen=True)
class RenderResult:
    text: str
    missing_headers: tuple[str, ...] = ()
    empty_fields: tuple[str, ...] = ()


@dataclass
class RecipientPreview:
    row_number: int
    email: str
    subject: str
    body: str
    values: dict[str, str]
    valid: bool
    problems: list[str] = field(default_factory=list)
    status: str = "Hazır"


@dataclass(frozen=True)
class SMTPConfig:
    host: str
    port: int
    security: str
    username: str
    password: str
    sender_address: str
    sender_name: str = ""
    reply_to: str = ""
    timeout: int = 30

    def validate(self) -> None:
        if not self.host.strip():
            raise MailBotError("SMTP sunucusu boş bırakılamaz.")
        if not 1 <= self.port <= 65535:
            raise MailBotError("SMTP portu 1 ile 65535 arasında olmalıdır.")
        if self.security not in {"STARTTLS", "SSL/TLS", "Yok"}:
            raise MailBotError("Bilinmeyen SMTP güvenlik seçimi.")
        if not is_valid_email(self.sender_address):
            raise MailBotError("Gönderen e-posta adresi geçerli değil.")
        for label, value in (
            ("Gönderen adı", self.sender_name),
            ("Yanıt adresi", self.reply_to),
            ("Kullanıcı adı", self.username),
        ):
            if NEWLINE_RE.search(value):
                raise MailBotError(f"{label} satır sonu karakteri içeremez.")
        if self.reply_to and not is_valid_email(self.reply_to):
            raise MailBotError("Yanıt adresi geçerli değil.")


def canonical_name(value: object) -> str:
    """Başlıkları Türkçe karakter ve büyük/küçük harf farkına dayanıklı eşleştirir."""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    text = text.translate(str.maketrans({"İ": "i", "I": "i", "ı": "i"})).casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def format_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, bool):
        return "Evet" if value else "Hayır"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _validate_headers(raw_headers: Sequence[object]) -> list[str]:
    last_used = -1
    for index, value in enumerate(raw_headers):
        if value is not None and str(value).strip():
            last_used = index
    if last_used < 0:
        raise MailBotError("Dosyanın ilk satırında sütun başlığı bulunamadı.")

    headers: list[str] = []
    seen: dict[str, str] = {}
    for index, value in enumerate(raw_headers[: last_used + 1], start=1):
        header = format_cell_value(value)
        if not header:
            raise MailBotError(f"{index}. sütunun başlığı boş. Lütfen ilk satırı düzeltin.")
        key = canonical_name(header)
        if key in seen:
            raise MailBotError(
                f"Birbiriyle aynı kabul edilen iki sütun var: '{seen[key]}' ve '{header}'."
            )
        seen[key] = header
        headers.append(header)
    return headers


def list_sheet_names(path: str | Path) -> list[str]:
    file_path = Path(path)
    if not file_path.exists():
        raise MailBotError("Seçilen veri dosyası bulunamadı.")
    if file_path.suffix.lower() == ".csv":
        return ["CSV"]
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise MailBotError("Yalnızca .xlsx, .xlsm ve .csv dosyaları desteklenir.")
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        names = list(workbook.sheetnames)
        workbook.close()
        return names
    except Exception as exc:  # openpyxl çeşitli dosya hataları döndürebilir
        raise MailBotError(f"Excel dosyası açılamadı: {exc}") from exc


def load_records(path: str | Path, sheet_name: str | None = None) -> tuple[list[str], list[Record]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _load_csv_records(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel_records(file_path, sheet_name)
    raise MailBotError("Yalnızca .xlsx, .xlsm ve .csv dosyaları desteklenir.")


def _load_excel_records(path: Path, sheet_name: str | None) -> tuple[list[str], list[Record]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        selected = sheet_name or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            workbook.close()
            raise MailBotError(f"'{selected}' adlı çalışma sayfası bulunamadı.")
        sheet = workbook[selected]
        iterator = sheet.iter_rows(values_only=True)
        first_row = next(iterator, None)
        if first_row is None:
            workbook.close()
            raise MailBotError("Seçilen çalışma sayfası boş.")
        headers = _validate_headers(first_row)
        records: list[Record] = []
        for row_number, row in enumerate(iterator, start=2):
            used = row[: len(headers)]
            if not any(value is not None and str(value).strip() for value in used):
                continue
            values = {
                header: format_cell_value(used[index] if index < len(used) else None)
                for index, header in enumerate(headers)
            }
            records.append(Record(row_number=row_number, values=values))
        workbook.close()
        return headers, records
    except MailBotError:
        raise
    except Exception as exc:
        raise MailBotError(f"Excel verileri okunamadı: {exc}") from exc


def _load_csv_records(path: Path) -> tuple[list[str], list[Record]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1254"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                rows = list(csv.reader(handle, dialect))
            if not rows:
                raise MailBotError("CSV dosyası boş.")
            headers = _validate_headers(rows[0])
            records: list[Record] = []
            for row_number, row in enumerate(rows[1:], start=2):
                used = row[: len(headers)]
                if not any(str(value).strip() for value in used):
                    continue
                values = {
                    header: format_cell_value(used[index] if index < len(used) else "")
                    for index, header in enumerate(headers)
                }
                records.append(Record(row_number=row_number, values=values))
            return headers, records
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except MailBotError:
            raise
        except Exception as exc:
            raise MailBotError(f"CSV verileri okunamadı: {exc}") from exc
    raise MailBotError(f"CSV karakter kodlaması okunamadı: {last_error}")


def placeholder_names(*templates: str) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for template in templates:
        protected = re.sub(r"\(\(([^()]*)\)\)", "", template)
        for match in PLACEHOLDER_RE.finditer(protected):
            name = match.group(1).strip()
            key = canonical_name(name)
            if key and key not in seen:
                seen.add(key)
                found.append(name)
    return found


def render_template(template: str, values: Mapping[str, object]) -> RenderResult:
    lookup = {canonical_name(key): format_cell_value(value) for key, value in values.items()}
    literals: list[str] = []

    def protect_literal(match: re.Match[str]) -> str:
        literals.append(f"({match.group(1)})")
        return f"\x00LITERAL{len(literals) - 1}\x00"

    protected = re.sub(r"\(\(([^()]*)\)\)", protect_literal, template)
    missing: list[str] = []
    empty: list[str] = []

    def replace(match: re.Match[str]) -> str:
        original = match.group(1).strip()
        key = canonical_name(original)
        if key not in lookup:
            if original not in missing:
                missing.append(original)
            return match.group(0)
        if not lookup[key] and original not in empty:
            empty.append(original)
        return lookup[key]

    rendered = PLACEHOLDER_RE.sub(replace, protected)
    for index, literal in enumerate(literals):
        rendered = rendered.replace(f"\x00LITERAL{index}\x00", literal)
    return RenderResult(rendered, tuple(missing), tuple(empty))


def is_valid_email(value: str) -> bool:
    return bool(value and not NEWLINE_RE.search(value) and EMAIL_RE.fullmatch(value.strip()))


def random_delay_seconds(min_seconds: float, max_seconds: float) -> float:
    """İki gönderim arasında, iki uç dahil olacak şekilde rastgele süre üretir."""
    return random.uniform(min_seconds, max_seconds)


def create_previews(
    records: Iterable[Record],
    email_column: str,
    subject_template: str,
    body_template: str,
    *,
    skip_duplicate_addresses: bool = True,
) -> list[RecipientPreview]:
    if not subject_template.strip():
        raise MailBotError("E-posta konusu boş bırakılamaz.")
    if not body_template.strip():
        raise MailBotError("E-posta metni boş bırakılamaz.")

    email_key = canonical_name(email_column)
    previews: list[RecipientPreview] = []
    seen_emails: set[str] = set()
    for record in records:
        value_lookup = {canonical_name(key): value for key, value in record.values.items()}
        email = value_lookup.get(email_key, "").strip()
        subject_result = render_template(subject_template, record.values)
        body_result = render_template(body_template, record.values)

        problems: list[str] = []
        missing = list(dict.fromkeys(subject_result.missing_headers + body_result.missing_headers))
        empty = list(dict.fromkeys(subject_result.empty_fields + body_result.empty_fields))
        if missing:
            problems.append("Excel'de olmayan alan: " + ", ".join(missing))
        if empty:
            problems.append("Boş alan: " + ", ".join(empty))
        if not is_valid_email(email):
            problems.append("Geçersiz e-posta")
        normalized_email = email.casefold()
        if skip_duplicate_addresses and normalized_email and normalized_email in seen_emails:
            problems.append("Tekrarlanan e-posta")
        if normalized_email:
            seen_emails.add(normalized_email)
        if NEWLINE_RE.search(subject_result.text):
            problems.append("Konu satır sonu içeriyor")

        previews.append(
            RecipientPreview(
                row_number=record.row_number,
                email=email,
                subject=subject_result.text,
                body=body_result.text,
                values=record.values,
                valid=not problems,
                problems=problems,
                status="Hazır" if not problems else "Atlanacak: " + "; ".join(problems),
            )
        )
    return previews


def build_message(
    config: SMTPConfig,
    recipient: str,
    subject: str,
    body: str,
    attachments: Sequence[str | Path] = (),
) -> EmailMessage:
    config.validate()
    if not is_valid_email(recipient):
        raise MailBotError(f"Geçersiz alıcı adresi: {recipient}")
    if NEWLINE_RE.search(subject):
        raise MailBotError("E-posta konusu satır sonu karakteri içeremez.")

    message = EmailMessage()
    message["From"] = formataddr((config.sender_name, config.sender_address))
    message["To"] = recipient
    message["Subject"] = subject
    if config.reply_to:
        message["Reply-To"] = config.reply_to
    message.set_content(body)

    for attachment in attachments:
        path = Path(attachment)
        if not path.is_file():
            raise MailBotError(f"Ek dosyası bulunamadı: {path}")
        mime_type, _ = mimetypes.guess_type(path.name)
        if mime_type:
            main_type, sub_type = mime_type.split("/", 1)
        else:
            main_type, sub_type = "application", "octet-stream"
        message.add_attachment(
            path.read_bytes(), maintype=main_type, subtype=sub_type, filename=path.name
        )
    return message


class SMTPMailer:
    def __init__(self, config: SMTPConfig):
        config.validate()
        self.config = config
        self.connection: smtplib.SMTP | smtplib.SMTP_SSL | None = None

    def __enter__(self) -> "SMTPMailer":
        context = ssl.create_default_context()
        try:
            if self.config.security == "SSL/TLS":
                self.connection = smtplib.SMTP_SSL(
                    self.config.host, self.config.port, timeout=self.config.timeout, context=context
                )
            else:
                self.connection = smtplib.SMTP(
                    self.config.host, self.config.port, timeout=self.config.timeout
                )
                self.connection.ehlo()
                if self.config.security == "STARTTLS":
                    self.connection.starttls(context=context)
                    self.connection.ehlo()
            if self.config.username:
                self._login()
            return self
        except Exception:
            self.close()
            raise

    def _login(self) -> None:
        if self.connection is None:
            raise MailBotError("SMTP bağlantısı açık değil.")
        try:
            self.config.username.encode("ascii")
            self.config.password.encode("ascii")
        except UnicodeEncodeError:
            self._login_utf8()
        else:
            self.connection.login(self.config.username, self.config.password)

    def _login_utf8(self) -> None:
        """Python smtplib'in ASCII sınırlamasını aşarak SASL PLAIN/LOGIN'i UTF-8 yollar."""
        if self.connection is None:
            raise MailBotError("SMTP bağlantısı açık değil.")
        advertised = self.connection.esmtp_features.get("auth", "")
        mechanisms = {item.upper() for item in advertised.split()}
        supported = [item for item in ("PLAIN", "LOGIN") if item in mechanisms]
        if not supported:
            raise MailBotError(
                "SMTP kullanıcı adı veya parola Türkçe/Unicode karakter içeriyor; "
                "sunucu UTF-8 ile kullanılabilen PLAIN/LOGIN doğrulaması sunmuyor. "
                "Tam e-posta adresinizi ve varsa sağlayıcınızın uygulama parolasını kullanın."
            )

        last_error: smtplib.SMTPAuthenticationError | None = None
        for mechanism in supported:
            if mechanism == "PLAIN":
                payload = f"\0{self.config.username}\0{self.config.password}".encode("utf-8")
                token = base64.b64encode(payload).decode("ascii")
                code, response = self.connection.docmd("AUTH", "PLAIN " + token)
            else:
                code, response = self.connection.docmd("AUTH", "LOGIN")
                if code == 334:
                    encoded_user = base64.b64encode(self.config.username.encode("utf-8")).decode("ascii")
                    code, response = self.connection.docmd(encoded_user)
                if code == 334:
                    encoded_password = base64.b64encode(self.config.password.encode("utf-8")).decode("ascii")
                    code, response = self.connection.docmd(encoded_password)

            if code in (235, 503):
                return
            last_error = smtplib.SMTPAuthenticationError(code, response)

        if last_error is not None:
            raise MailBotError(
                "SMTP kimlik doğrulaması başarısız. Kullanıcı adı/parolanız Türkçe veya "
                "Unicode karakter içeriyor. Gmail gibi bir sağlayıcı kullanıyorsanız normal "
                "hesap parolası yerine uygulama parolası girmeniz gerekebilir."
            ) from last_error
        raise MailBotError("SMTP kimlik doğrulaması tamamlanamadı.")

    def send(
        self,
        recipient: str,
        subject: str,
        body: str,
        attachments: Sequence[str | Path] = (),
    ) -> None:
        if self.connection is None:
            raise MailBotError("SMTP bağlantısı açık değil.")
        message = build_message(self.config, recipient, subject, body, attachments)
        self.connection.send_message(message)

    def close(self) -> None:
        if self.connection is not None:
            try:
                self.connection.quit()
            except Exception:
                try:
                    self.connection.close()
                except Exception:
                    pass
            finally:
                self.connection = None

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        self.close()


def smtp_error_text(exc: Exception) -> str:
    if isinstance(exc, UnicodeEncodeError):
        return (
            "SMTP kullanıcı adı veya parola Türkçe/Unicode karakter içeriyor. "
            "Kullanıcı adı olarak tam e-posta adresini, parola olarak sağlayıcınızın "
            "uygulama parolasını kullanmayı deneyin."
        )
    if isinstance(exc, smtplib.SMTPAuthenticationError):
        return "SMTP kimlik doğrulaması başarısız. Kullanıcı adı ve uygulama parolasını kontrol edin."
    if isinstance(exc, smtplib.SMTPConnectError):
        return "SMTP sunucusuna bağlantı kurulamadı."
    if isinstance(exc, smtplib.SMTPRecipientsRefused):
        return "Alıcı adresi posta sunucusu tarafından reddedildi."
    if isinstance(exc, smtplib.SMTPSenderRefused):
        return "Gönderen adresi posta sunucusu tarafından reddedildi."
    if isinstance(exc, (TimeoutError, OSError)):
        return f"Ağ/bağlantı hatası: {exc}"
    return str(exc) or exc.__class__.__name__


def write_log_row(
    log_path: str | Path,
    *,
    row_number: int,
    email: str,
    subject: str,
    result: str,
    detail: str = "",
) -> None:
    path = Path(log_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    new_file = not path.exists() or path.stat().st_size == 0
    with path.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        if new_file:
            writer.writerow(["zaman", "excel_satiri", "eposta", "konu", "sonuc", "detay"])
        writer.writerow(
            [
                datetime.now().isoformat(timespec="seconds"),
                row_number,
                email,
                subject,
                result,
                detail,
            ]
        )


def prepare_log_file(log_path: str | Path) -> None:
    """İlk e-posta gönderilmeden önce günlük hedefinin yazılabilir olduğunu doğrular."""
    path = Path(log_path)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("x", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle, delimiter=";")
            writer.writerow(["zaman", "excel_satiri", "eposta", "konu", "sonuc", "detay"])
    except FileExistsError:
        if not path.is_file():
            raise MailBotError("Gönderim günlüğü hedefi bir dosya değil.")
    except OSError as exc:
        raise MailBotError(f"Gönderim günlüğü oluşturulamadı: {exc}") from exc
